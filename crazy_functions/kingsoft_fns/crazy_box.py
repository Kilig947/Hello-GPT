#! .\venv\
# encoding: utf-8
# @Time   : 2023/6/14
# @Author : Spike
# @Descr   :
import os
import json
import requests
import re
import time
import xmindparser
from typing import Dict
import typing as typing
from comm_tools import func_box, ocr_tools, toolbox, database_processor, Langchain_cn
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from crazy_functions import crazy_utils
from request_llm import bridge_all
from crazy_functions.kingsoft_fns import crzay_kingsoft


class Utils:

    def __init__(self):
        self.find_keys_type = 'type'
        self.find_picture_source = ['caption', 'imgID', 'sourceKey']
        self.find_document_source = ['wpsDocumentLink', 'wpsDocumentName', 'wpsDocumentType']
        self.find_document_tags = ['WPSDocument']
        self.find_picture_tags = ['picture', 'processon']
        self.picture_format = ['.jpeg', '.jpg', '.png', '.gif', '.bmp', '.tiff']
        self.comments = []

    def find_all_text_keys(self, dictionary, parent_type=None, text_values=None, filter_type=''):
        """
        Args:
            dictionary: 字典或列表
            parent_type: 匹配的type，作为新列表的key，用于分类
            text_values: 存储列表
            filter_type: 当前层级find_keys_type==filter_type时，不继续往下嵌套
        Returns:
            text_values和排序后的context_
        """
        # 初始化 text_values 为空列表，用于存储找到的所有text值
        if text_values is None:
            text_values = []
        # 如果输入的dictionary不是字典类型，返回已收集到的text值
        if not isinstance(dictionary, dict):
            return text_values
        # 获取当前层级的 type 值
        current_type = dictionary.get(self.find_keys_type, parent_type)
        # 如果字典中包含 'text' 或 'caption' 键，将对应的值添加到 text_values 列表中
        if 'comments' in dictionary:
            temp = dictionary.get('comments', [])
            for t in temp:
                if type(t) is dict:
                    self.comments.append(t.get('key'))
        if 'text' in dictionary:
            content_value = dictionary.get('text', None)
            text_values.append({current_type: content_value})
        if 'caption' in dictionary:
            temp = {}
            for key in self.find_picture_source:
                temp[key] = dictionary.get(key, None)
            text_values.append({current_type: temp})
        if 'wpsDocumentId' in dictionary:
            temp = {}
            for key in self.find_document_source:
                temp[key] = dictionary.get(key, None)
            text_values.append({current_type: temp})
        # 如果当前类型不等于 filter_type，则继续遍历子级属性
        if current_type != filter_type:
            for key, value in dictionary.items():
                if isinstance(value, dict):
                    self.find_all_text_keys(value, current_type, text_values, filter_type)
                elif isinstance(value, list):
                    for item in value:
                        if isinstance(item, dict):
                            self.find_all_text_keys(item, current_type, text_values, filter_type)
        return text_values

    def statistical_results(self, text_values, img_proce=False):
        """
        Args: 提取爬虫内嵌图片、文件等等信息
            text_values: dict
            img_proce: 图片标识
        Returns: （元数据， 组合数据， 图片dict， 文件dict）
        """
        context_ = []
        pic_dict = {}
        file_dict = {}
        for i in text_values:
            for key, value in i.items():
                if key in self.find_picture_tags:
                    if img_proce:
                        mark = f'{key}"""\n{value["sourceKey"]}\n"""\n'
                        if value["caption"]: mark += f'\n{key}:{value["caption"]}\n\n'
                        context_.append(mark)
                        pic_dict[value['sourceKey']] = value['imgID']
                    else:
                        if value["caption"]: context_.append(f'{key}描述: {value["caption"]}\n')
                        pic_dict[value['sourceKey']] = value['imgID']
                elif key in self.find_document_tags:
                    mark = f'{value["wpsDocumentName"]}: {value["wpsDocumentLink"]}'
                    file_dict.update({value["wpsDocumentName"]: value["wpsDocumentLink"]})
                    context_.append(mark)
                else:
                    context_.append(value)
        context_ = '\n'.join(context_)
        return text_values, context_, pic_dict, file_dict

    def write_markdown(self, data, hosts, file_name):
        """
        Args: 将data写入md文件
            data: 数据
            hosts: 用户标识
            file_name: 另取文件名
        Returns: 写入的文件地址
        """
        user_path = os.path.join(func_box.users_path, hosts, 'markdown')
        os.makedirs(user_path, exist_ok=True)
        md_file = os.path.join(user_path, f"{file_name}.md")
        with open(file=md_file, mode='w') as f:
            f.write(data)
        return md_file

    def md_2_html(self, data, hosts, file_name):
        """
        Args: 将data写入md文件
            data: 数据
            hosts: 用户标识
            file_name: 另取文件名
        Returns: 写入的文件地址
        """
        data = toolbox.markdown_convertion(data)
        user_path = os.path.join(func_box.users_path, hosts, 'view_html')
        os.makedirs(user_path, exist_ok=True)
        md_file = os.path.join(user_path, f"{file_name}.html")
        with open(file=md_file, mode='w') as f:
            f.write(data)
        return md_file

    def markdown_to_flow_chart(self, data, hosts, file_name):
        """
        Args: 调用markmap-cli
            data: 要写入md的数据
            hosts: 用户标识
            file_name: 要写入的文件名
        Returns: [md, 流程图] 文件
        """
        user_path = os.path.join(func_box.users_path, hosts, 'mark_map')
        os.makedirs(user_path, exist_ok=True)
        md_file = self.write_markdown(data, hosts, file_name)
        html_file = os.path.join(user_path, f"{file_name}.html")
        func_box.Shell(f'npx markmap-cli --no-open "{md_file}" -o "{html_file}"').read()
        return md_file, html_file

    def split_startswith_txt(self, link_limit, start='http', domain_name: list = ['']):
        link = str(link_limit).split()
        links = []
        for i in link:
            if i.startswith(start) and any(name in i for name in domain_name):
                links.append(i)
        return links

    def global_search_for_files(self, file_path, matching: list):
        file_list = []
        if os.path.isfile(file_path):
            file_list.append(file_path)
        for root, dirs, files in os.walk(file_path):
            for file in files:
                for math in matching:
                    if str(math).lower() in str(file).lower():
                        file_list.append(os.path.join(root, file))
        return file_list


class ExcelHandle:

    def __init__(self, ipaddr='temp', temp_file='', sheet='测试要点'):
        self.user_path = os.path.join(func_box.base_path, 'private_upload', ipaddr, 'test_case',
                                      func_box.created_atime())
        os.makedirs(f'{self.user_path}', exist_ok=True)
        if not temp_file:
            self.template_excel = os.path.join(func_box.base_path, 'docs/template/客户端测试用例模版.xlsx')
        if os.path.exists(temp_file):
            self.template_excel = temp_file
        elif temp_file.startswith('http'):
            self.template_excel = \
            crzay_kingsoft.get_kdocs_files(temp_file, project_folder=self.user_path, type='xlsx', ipaddr=ipaddr)[0]
        else:
            self.template_excel = os.path.join(func_box.base_path, 'docs/template/客户端测试用例模版.xlsx')
        if not self.template_excel:
            self.template_excel = os.path.join(func_box.base_path, 'docs/template/客户端测试用例模版.xlsx')
        self.workbook = load_workbook(self.template_excel)
        self.sheet = sheet
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.green_fill = PatternFill(start_color="1abc9c", end_color="1abc9c", fill_type="solid")
        self.red_fill = PatternFill(start_color="ff7f50", end_color="ff7f50", fill_type="solid")
        # 定义边框样式
        border_style = Side(style='thin', color="000000")
        self.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        if str(self.sheet) not in self.workbook.sheetnames:
            self.sheet = self.workbook.active.title


    def lpvoid_lpbuffe(self, data_list: list, filename=''):
        # 加载现有的 Excel 文件        # 选择要操作的工作表, 默认是测试要点
        if self.sheet in self.workbook.sheetnames:
            worksheet = self.workbook[self.sheet]
        else:
            worksheet = self.workbook.create_sheet(self.sheet)
        # 定义起始行号
        start_row = find_index_inlist(self.read_as_dict()[self.sheet], ['操作步骤', '前置条件', '预期结果']) + 2
        # 创建一个黄色的填充样式
        # 遍历数据列表
        for row_data in data_list:
            # 写入每一行的数据到指定的单元格范围
            for col_num, value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=start_row, column=col_num)
                try:
                    cell.value = str(value).strip()
                    cell.border = self.border
                    cell.alignment = Alignment(horizontal='center', vertical='center',
                                               wrapText=True)  # 设置水平和垂直方向均居中对齐，并自动换行
                    # 判断 value 是否为 '插件补充的用例'
                    if '插件补充的用例' in str(value):
                        cell.fill = self.yellow_fill
                    font = Font(name='苹方-简', size=11)
                    cell.font = font
                except Exception:
                    print(row_data, value)
                    func_box.通知机器人(error=f'写入excel错误啦\n\n```\n\n{row_data}\n\n{value}\n\n```'
                                              f'\n\n```\n\n{toolbox.trimmed_format_exc()}```\n\n')
            # 增加起始行号
            start_row += 1
        merge_cell, = toolbox.get_conf('merge_cell')
        if merge_cell: self.merge_same_cells()  # 还原被拆分的合并单元格
        # 保存 Excel 文件
        time_stamp = time.strftime("%Y-%m-%d-%H", time.localtime())
        if filename == '':
            filename = time.strftime("%Y-%m-%d-%H", time.localtime()) + '_temp'
        else:
            f"{time_stamp}_{filename}"
        test_case_path = f'{os.path.join(self.user_path, filename)}.xlsx'
        # 遇到文件无法保存时，再拆开图片
        try:
            self.workbook.save(test_case_path)
        except Exception as f:
            test_case_path = self.template_excel
        return test_case_path

    def read_as_dict(self, only_sheet=True):
        data_dict = {}
        # 遍历每个工作表
        if only_sheet:
            sheet_list = [self.sheet]
        else:
            sheet_list = self.workbook.sheetnames
        for sheet_name in sheet_list:
            sheet = self.workbook[sheet_name]
            sheet_data = []
            # 遍历每一行
            sheet_temp_count = 0
            for row in sheet.iter_rows(values_only=True):
                # 过滤尾部的空行
                row = tuple(x for x in row if x is not None and x != row[-1])
                if row:
                    sheet_data.append(row)
                else:
                    sheet_temp_count += 1
                if sheet_temp_count >= 20: break
            # 将工作表名作为字典的键，行数据作为值
            data_dict[sheet_name] = sheet_data
        return data_dict

    def split_merged_cells(self):
        # 加载Excel文件
        ws = self.workbook[self.sheet]
        # 获取合并单元格的范围
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            # 获取合并单元格的起始行、起始列、结束行、结束列
            start_row = merged_range.min_row
            start_col = merged_range.min_col
            end_row = merged_range.max_row
            end_col = merged_range.max_col
            # 获取合并单元格的值
            value = ws.cell(start_row, start_col).value
            # 拆分合并单元格
            ws.unmerge_cells(str(merged_range))
            # 在每个拆分后的单元格中填入值
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row, col)
                    cell.value = value
        # 保存结果
        self.workbook.save(self.template_excel)

    def merge_same_cells(self, truncation=10):
        # 加载xlsx文件
        ws = self.workbook[self.sheet]
        # 遍历每个单元格（列优先遍历）
        column_counter = {'row': 0, 'col': 0}
        for col_index in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_index)
            row_start = None
            last_column_empty = True
            for row_index in range(1, ws.max_row + 1):
                current_cell = ws[f"{col_letter}{row_index}"]
                next_cell = ws[f"{col_letter}{row_index + 1}"]
                # 当前单元格与下个单元格内容相同时，都不为空，并记录合并范围row_start
                if row_start is None and current_cell.value == next_cell.value and current_cell.value is not None:
                    row_start = row_index
                # 当前单元格与下个单元格内容不同时或任何一个为空时，记录合并范围row_end，并执行合并
                elif row_start is not None and (
                        current_cell.value != next_cell.value or current_cell.value is None or next_cell.value is None):
                    row_end = row_index
                    ws.merge_cells(f"{col_letter}{row_start}:{col_letter}{row_end}")
                    row_start = None
                # # 设置边框样式
                current_cell.border = self.border
                next_cell.border = self.border
                # 当列超过10行为空，跳出循环
                if not current_cell.value:
                    column_counter['row'] += 1
                    if column_counter['row'] > truncation:
                        column_counter['row'] = 0
                        break
                # 检查当前列是否为空
            if all(cell.value is None for cell in ws[col_letter]):
                if last_column_empty:  # 如果上一列也为空，增加计数器
                    column_counter['col'] += 1
                    if column_counter['col'] > truncation:  # 如果空列超过所设定的上限，跳出循环
                        break
                else:  # 如果上一列非空，重置计数器
                    column_counter['col'] = 1
                    last_column_empty = True
            else:  # 如果当前列非空，重置计数器和 last_column_empty 标记
                last_column_empty = False
                column_counter['col'] = 0
        self.workbook.save(self.template_excel)


class XmindHandle():

    def __int__(self):
        pass

    def _WalkTopic(self, dictXmind: Dict, resultDict: Dict):
        strTitle: typing.AnyStr = dictXmind['title']
        if 'topics' in dictXmind:
            pass
            # print(dictXmind['topics'])
            listTopics: typing.List = dictXmind['topics']

            if (listTopics.__len__() > 0):
                resultDict[strTitle] = {}
                for topic in listTopics:
                    self._WalkTopic(topic, resultDict[strTitle])
        else:
            resultDict[strTitle] = strTitle

    def _Print2MDList(self, dictOri: typing.Dict) -> typing.AnyStr:
        levelOri = 0
        listStr = []

        def Print2MDListInternal(dictContent: typing.Dict, level):
            if type(dictContent).__name__ != 'dict':
                return
            level = level + 1
            for topic, topicDict in dictContent.items():
                listStr.append('  ' * (level - 1))
                listStr.append('- ')
                listStr.append(topic)
                listStr.append('\n')
                Print2MDListInternal(topicDict, level)

        Print2MDListInternal(dictOri, levelOri)
        return ''.join(listStr)

    def xmind_2_md(self, pathSource):
        try:
            dictSheet = xmindparser.xmind_to_dict(pathSource)

        except:
            import xmind
            workbook = xmind.load(pathSource)
            sheet = workbook.getPrimarySheet()
            dictSheet = [sheet.getData()]
        dictResult: Dict = {}
        xm_content = ''
        md_path = []
        for canvas in dictSheet:
            self._WalkTopic(canvas['topic'], dictResult)
            strResult = self._Print2MDList(dictResult)
            xm_content += strResult
            temp_path = os.path.join(os.path.dirname(os.path.dirname(pathSource)), 'markdown')
            os.makedirs(temp_path, exist_ok=True)
            pathOutput = os.path.join(temp_path, f'{os.path.basename(pathSource)}_{canvas["title"]}.md')
            with open(pathOutput, 'w', encoding='utf-8') as f:
                f.write(strResult)
                md_path.append(pathOutput)
        return xm_content, md_path


# <---------------------------------------乱七八糟的方法，有点用，很好用----------------------------------------->
def find_index_inlist(data_list: list, search_terms: list) -> int:
    """ 在data_list找到符合search_terms字符串，找到后直接返回下标
    Args:
        data_list: list数据，最多往里面找两层
        search_terms: list数据，符合一个就返回数据
    Returns: 对应的下标
    """
    for i, sublist in enumerate(data_list):
        if any(term in str(sublist) for term in search_terms):
            return i
        for j, item in enumerate(sublist):
            if any(term in str(item) for term in search_terms):
                return i
    return 0  # 如果没有找到匹配的元素，则返回初始坐标


def file_extraction_intype(files, file_types, file_limit, chatbot, history, llm_kwargs, plugin_kwargs):
    """
    Args:
        file_routing: 文件路径
        file_limit:  存储解析后的文档list
        chatbot: 对话组件
    Returns: None
    """
    # 文件读取
    file_routing = []
    if type(file_types) is dict: files = [files[f] for f in files]
    for t in file_types:
        for f in files:
            _, routing, _ = crazy_utils.get_files_from_everything(f, t, ipaddr=llm_kwargs['ipaddr'])
            file_routing.extend(routing)
    for file_path in file_routing:
        chatbot.append(['检查文件是否符合格式要求，并解析文件', f'`{file_path.replace(func_box.base_path, ".")}`' +
                        f"\t...正在解析本地文件\n\n"])
        yield from toolbox.update_ui(chatbot, history)
        title = long_name_processing(os.path.basename(file_path))
        if file_path.endswith('pdf'):
            file_content, _ = crazy_utils.read_and_clean_pdf_text(file_path)
            content = "".join(file_content)
            file_limit.extend([title, content])
        elif file_path.endswith('docx') or file_path.endswith('doc'):
            pass
        elif file_path.endswith('xmind'):
            file_content, _path = XmindHandle().xmind_2_md(pathSource=file_path)
            file_limit.extend([title, file_content])
        elif file_path.endswith('xlsx') or file_path.endswith('xls'):
            sheet, = json_args_return(plugin_kwargs, keys=['读取指定Sheet'], default='测试要点')
            # 创建文件对象
            ex_handle = ExcelHandle(temp_file=file_path, sheet=sheet)
            if sheet in ex_handle.workbook.sheetnames:
                ex_handle.split_merged_cells()
                xlsx_dict = ex_handle.read_as_dict()
                file_content = xlsx_dict.get(sheet)
                file_limit.extend([title, file_content])
            else:
                active_sheet = ex_handle.workbook.active.title
                ex_handle.sheet = active_sheet
                ex_handle.split_merged_cells()
                xlsx_dict = ex_handle.read_as_dict()
                active_content = xlsx_dict.get(active_sheet)
                file_limit.extend([title, active_content])
                chatbot.append(['可以开始了么？',
                                f'无法在`{os.path.basename(file_path)}`找到`{sheet}`工作表'
                                f'将读取上次预览的活动工作表`{active_sheet}`，'
                                f'若你的用例工作表是其他名称, 请及时暂停插件运行，并在自定义插件配置中更改'
                                f'{func_box.html_tag_color("读取指定Sheet")}。'])
            plugin_kwargs['写入指定模版'] = file_path
            plugin_kwargs['写入指定Sheet'] = ex_handle.sheet
            yield from toolbox.update_ui(chatbot, history)
        else:
            with open(file_path, mode='r') as f:
                file_content = f.read()
                file_limit.extend([title, file_content])
        yield from toolbox.update_ui(chatbot, history)


def json_args_return(kwargs, keys: list, default=None) -> list:
    """
    Args: 提取插件的调优参数，如果有，则返回取到的值，如果无，则返回False
        kwargs: 一般是plugin_kwargs
        keys: 需要取得key
        default: 找不到时总得返回什么东西
    Returns: 有key返value，无key返False
    """
    temp = [default for i in range(len(keys))]
    for i in range(len(keys)):
        try:
            temp[i] = kwargs[keys[i]]
        except Exception:
            try:
                temp[i] = json.loads(kwargs['advanced_arg'])[keys[i]]
            except Exception as f:
                try:
                    temp[i] = kwargs['parameters_def'][keys[i]]
                except Exception as f:
                    temp[i] = default
    return temp


def long_name_processing(file_name):
    """
    Args:
        file_name: 文件名取材，如果是list，则取下标0，转换为str， 如果是str则取最多20个字符
    Returns: 返回处理过的文件名
    """
    if type(file_name) is list:
        file_name = file_name[0]
    if len(file_name) > 20:
        temp = file_name.splitlines()
        for i in temp:
            if i:
                file_name = func_box.replace_special_chars(i)
                break
    if file_name.find('.') != -1:
        file_name = "".join(file_name.split('.')[:-1])
    return file_name


# <---------------------------------------插件用了都说好方法----------------------------------------->
def split_list_token_limit(data, get_num, max_num=500):
    header_index = find_index_inlist(data_list=data, search_terms=['操作步骤', '前置条件', '预期结果'])
    header_data = data[header_index]
    max_num -= len(str(header_data))
    temp_list = []
    split_data = []
    for index in data[header_index + 1:]:
        if get_num(str(temp_list)) > max_num:
            temp_list.insert(0, header_data)
            split_data.append(json.dumps(temp_list, ensure_ascii=False))
            temp_list = []
        else:
            temp_list.append(index)
    return split_data


def split_content_limit(inputs: str, llm_kwargs, chatbot, history) -> list:
    """
    Args:
        inputs: 需要提取拆分的提问信息
        llm_kwargs: 调优参数
        chatbot: 对话组件
        history: 历史记录
    Returns: [拆分1， 拆分2]
    """
    model = llm_kwargs['llm_model']
    if model.find('&') != -1:  # 判断是否多模型，如果多模型，那么使用tokens最少的进行拆分
        models = str(model).split('&')
        _tokens = []
        _num_func = {}
        for _model in models:
            num_s = bridge_all.model_info[_model]['max_token']
            _tokens.append(num_s)
            _num_func[num_s] = _model
        all_tokens = min(_tokens)
        get_token_num = bridge_all.model_info[_num_func[all_tokens]]['token_cnt']
    else:
        all_tokens = bridge_all.model_info[model]['max_token']
        get_token_num = bridge_all.model_info[model]['token_cnt']
    max_token = all_tokens / 2 - all_tokens / 4  # 考虑到对话+回答会超过tokens,所以/2
    segments = []
    if type(inputs) is list:
        if get_token_num(str(inputs)) > max_token:
            chatbot.append([None,
                            f'{func_box.html_tag_color(inputs[0][:10])}...对话数据预计会超出{all_tokens}tokens限制, 拆分中...'])
            segments.extend(split_list_token_limit(data=inputs, get_num=get_token_num, max_num=max_token))
        else:
            segments.append(json.dumps(inputs, ensure_ascii=False))
    else:
        inputs = inputs.split('\n---\n')
        for input_ in inputs:
            if get_token_num(input_) > max_token:
                chatbot.append([None,
                                f'{func_box.html_tag_color(input_[:10])}...对话数据预计会超出{all_tokens}tokens限制, 拆分中...'])
                yield from toolbox.update_ui(chatbot, history)
                segments.extend(crazy_utils.breakdown_txt_to_satisfy_token_limit_for_pdf(input_, get_token_num, max_token))
            else:
                segments.append(input_)
    yield from toolbox.update_ui(chatbot, history)
    return segments


def input_output_processing(gpt_response_collection, llm_kwargs, plugin_kwargs, chatbot, history,
                            default_prompt: str = False, knowledge_base: bool = False, task_tag=''):
    """
    Args:
        gpt_response_collection:  多线程GPT的返回结果or文件读取处理后的结果
        plugin_kwargs: 对话使用的插件参数
        chatbot: 对话组件
        history: 历史对话
        llm_kwargs:  调优参数
        default_prompt: 默认Prompt, 如果为False，则不添加提示词
        knowledge_base: 是否启用知识库
    Returns: 下次使用？
        inputs_array， inputs_show_user_array
    """
    inputs_array = []
    inputs_show_user_array = []
    kwargs_prompt, prompt_cls = json_args_return(plugin_kwargs, ['预期产出提示词', '提示词分类'])
    if not prompt_cls or prompt_cls == '个人':  # 当提示词分类获取不到或个人时，使用个人prompt
        prompt_cls_tab = f'prompt_{llm_kwargs["ipaddr"]}'
    else:
        prompt_cls_tab = f'prompt_{prompt_cls}_sys'
    if default_prompt: kwargs_prompt = default_prompt
    chatbot.append([None, f'接下来使用的Prompt是`{prompt_cls}`分类下的：`{kwargs_prompt}`'
                          f', 你可以在{func_box.html_tag_color("自定义插件参数")}中指定另一个Prompt哦～'])
    time.sleep(1)
    prompt = database_processor.SqliteHandle(table=prompt_cls_tab).find_prompt_result(kwargs_prompt)
    for inputs, you_say in zip(gpt_response_collection[1::2], gpt_response_collection[0::2]):
        content_limit = yield from split_content_limit(inputs, llm_kwargs, chatbot, history)
        try:
            plugin_kwargs['上阶段文件'] = you_say
            plugin_kwargs[you_say] = {}
            plugin_kwargs[you_say]['原测试用例数据'] = [json.loads(limit)[1:] for limit in content_limit]
            plugin_kwargs[you_say]['原测试用例表头'] = json.loads(content_limit[0])[0]
        except Exception as f:
            print(f'读取原测试用例报错 {f}')
        for limit in content_limit:
            if knowledge_base:
                try:
                    limit = yield from Langchain_cn.knowledge_base_query(limit, chatbot, history, llm_kwargs,
                                                                         plugin_kwargs)
                except Exception as f:
                    func_box.通知机器人(f'读取知识库失败，请检查{f}')
            inputs_array.append(func_box.replace_expected_text(prompt, content=limit, expect='{{{v}}}'))
            inputs_show_user_array.append(you_say + task_tag)
    yield from toolbox.update_ui(chatbot, history)
    return inputs_array, inputs_show_user_array


def submit_multithreaded_tasks(inputs_array, inputs_show_user_array, llm_kwargs, chatbot, history, plugin_kwargs):
    """
    Args: 提交多线程任务
        inputs_array: 需要提交给gpt的任务列表
        inputs_show_user_array: 显示在user页面上信息
        llm_kwargs: 调优参数
        chatbot: 对话组件
        history: 历史对话
        plugin_kwargs: 插件调优参数
    Returns:  将对话结果返回[输入, 输出]
    """
    if len(inputs_array) == 1:
        # 折叠输出
        # if len(inputs_array[0]) > 200:
        #     inputs_show_user = inputs_array[0][:100]+f"\n\n{func_box.html_tag_color('......超过200个字符折叠......')}\n\n"+inputs_array[0][-100:]
        # else:
        inputs_show_user = inputs_array[0]
        gpt_say = yield from crazy_utils.request_gpt_model_in_new_thread_with_ui_alive(
            inputs=inputs_array[0], inputs_show_user=inputs_show_user,
            llm_kwargs=llm_kwargs, chatbot=chatbot, history=[],
            sys_prompt="", refresh_interval=0.1
        )
        gpt_response_collection = [inputs_show_user_array[0], gpt_say]
        history.extend(gpt_response_collection)
    else:
        gpt_response_collection = yield from crazy_utils.request_gpt_model_multi_threads_with_very_awesome_ui_and_high_efficiency(
            inputs_array=inputs_array,
            inputs_show_user_array=inputs_show_user_array,
            llm_kwargs=llm_kwargs,
            chatbot=chatbot,
            history_array=[[""] for _ in range(len(inputs_array))],
            sys_prompt_array=["" for _ in range(len(inputs_array))],
            # max_workers=5,  # OpenAI所允许的最大并行过载
            scroller_max_len=80,
        )
        # 是否展示任务结果
        kwargs_is_show, = json_args_return(plugin_kwargs, ['显示过程'])
        if kwargs_is_show:
            for results in list(zip(inputs_array, gpt_response_collection[1::2])):
                chatbot.append(results)
                history.extend(results)
                yield from toolbox.update_ui(chatbot, history)
    return gpt_response_collection


def func_拆分与提问(file_limit, llm_kwargs, plugin_kwargs, chatbot, history, args_keys: list, task_tag: str = ''):
    if args_keys[1]:
        plugin_kwargs['关联知识库'] = args_keys[1]
    multi_model_parallelism, = json_args_return(plugin_kwargs, ['多模型并行'], llm_kwargs['llm_model'])
    llm_kwargs['llm_model'] = multi_model_parallelism
    split_content_limit = yield from input_output_processing(file_limit, llm_kwargs, plugin_kwargs,
                                                             chatbot, history, default_prompt=args_keys[0],
                                                             knowledge_base=args_keys[1], task_tag=task_tag)
    inputs_array, inputs_show_user_array = split_content_limit
    gpt_response_collection = yield from submit_multithreaded_tasks(inputs_array, inputs_show_user_array,
                                                                    llm_kwargs, chatbot, history,
                                                                    plugin_kwargs)
    return gpt_response_collection


# <---------------------------------------写入文件方法----------------------------------------->
def file_classification_to_dict(gpt_response_collection):
    """
    接收gpt多线程的返回数据，将输入相同的作为key, gpt返回以列表形式添加到对应的key中，主要是为了区分不用文件的输入
    Args:
        gpt_response_collection: 多线程GPT的返回耶
    Returns: {'文件': [结果1， 结果2...], '文件2': [结果1， 结果2...]}
    """
    file_classification = {}
    for inputs, you_say in zip(gpt_response_collection[1::2], gpt_response_collection[0::2]):
        file_classification[you_say] = []
    for inputs, you_say in zip(gpt_response_collection[1::2], gpt_response_collection[0::2]):
        file_classification[you_say].append(inputs)
    return file_classification


def batch_recognition_images_to_md(img_list, ipaddr):
    """
    Args: 将图片批量识别然后写入md文件
        img_list: 图片地址list
        ipaddr: 用户所属标识
    Returns: [文件list]
    """
    temp_list = []
    for img in img_list:
        if os.path.exists(img):
            img_content, img_result, _ = ocr_tools.Paddle_ocr_select(ipaddr=ipaddr, trust_value=True
                                                                     ).img_def_content(img_path=img)
            temp_file = os.path.join(func_box.users_path, ipaddr, 'ocr_to_md', img_content.splitlines()[0][:20] + '.md')
            with open(temp_file, mode='w') as f:
                f.write(f"{func_box.html_view_blank(temp_file)}\n\n" + img_content)
            temp_list.append(temp_list)
        else:
            print(img, '文件路径不存在')
    return temp_list


def name_de_add_sort(response, index=0):
    unique_tuples = set(tuple(lst) for lst in response)
    de_result = [list(tpl) for tpl in unique_tuples]
    d = {}
    for i, v in enumerate(de_result):
        if len(v) >= index:
            if v[index] not in d:
                d[v[index]] = i
        else:
            d[v[len(v)]] = i
    de_result.sort(key=lambda x: d[x[index]])
    return de_result


def parsing_json_in_text(txt_data: list, old_case, filter_list: list = 'None----', tags='插件补充的用例', sort_index=0):
    response = []
    desc = '\n\n---\n\n'.join(txt_data)
    for index in range(len(old_case)):
        supplementary_data = []
        content_data = txt_data[index]
        if 'raise ConnectionAbortedError jsokf' in content_data:
            # 尝试修复超出Token限制导致的Json数据结构错误
            content_data = "\n".join([item for item in str(content_data).splitlines() if item != ''][:-1])
            if re.search(r'[^\w\s\]]', content_data[-1]):  # 判断是不是有,号之类的特殊字符
                content_data = content_data[:-1]  # 有则排除
            content_data += ']'
        # 尝试补充一些错误的JSON数据
        fix_data = content_data.replace('][', '],[').replace(']\n[', '],[')
        fix_data = fix_data.replace('\n...\n', '').replace('\n\n...\n\n', '')
        pattern = r'\[[^\[\]]*\]'
        result = re.findall(pattern, fix_data)
        for sp in result:
            __list = []
            try:
                __list = json.loads(sp)
                supplementary_data.append(__list)
            except:
                func_box.通知机器人(f'{sp} 测试用例转dict失败了来看看')
        if len(txt_data) != len(old_case): index = -1  # 兼容一下哈
        # 过滤掉产出带的表头数据
        filtered_supplementary_data = [data for data in supplementary_data
                                     if filter_list[:5] != data[:5] or filter_list[-5:] != data[-5:]]
        # 检查 filtered_supplementary_data 是否为空
        if not filtered_supplementary_data:
            max_length = 0  # 或其他合适的默认值
        else:
            max_length = max(len(lst) for lst in filtered_supplementary_data)
        supplement_temp_data = [lst + [''] * (max_length - len(lst)) for lst in filtered_supplementary_data]
        for new_case in supplement_temp_data:
            if new_case not in old_case[index] and new_case + [tags] not in old_case[index]:
                old_case[index].append(new_case + [tags])
        response.extend(old_case[index])
    # 按照名称排列重组
    response = name_de_add_sort(response, sort_index)
    return response, desc


def write_test_cases(gpt_response_collection, llm_kwargs, plugin_kwargs, chatbot, history):
    """
    Args:
        gpt_response_collection: [输入文件标题， 输出]
        llm_kwargs: 调优参数
        plugin_kwargs: 插件调优参数
        chatbot: 对话组件
        history: 对话历史
        file_key: 存入历史文件
    Returns: None
    """
    template_file, sheet, sort_index = json_args_return(plugin_kwargs, ['写入指定模版', '写入指定Sheet', '用例下标排序'])
    file_classification = file_classification_to_dict(gpt_response_collection)
    chat_file_list = ''
    you_say = '准备将测试用例写入Excel中...'
    chatbot.append([you_say, chat_file_list])
    yield from toolbox.update_ui(chatbot, history)
    files_limit = []
    for file_name in file_classification:
        test_case = []
        for value in file_classification[file_name]:
            test_case_content = value.splitlines()
            for i in test_case_content:
                if re.findall(r'\|\s*[:|-]+\s*\|', i):  # 过滤表头
                    test_case = test_case[:-1]
                    continue
                if i.find('|') != -1:
                    test_case.append([func_box.clean_br_string(i) for i in i.split('|')[1:]])
                elif i.find('｜') != -1:
                    test_case.append([func_box.clean_br_string(i) for i in i.split('｜')[1:]])
                else:
                    print('脏数据过滤，这个不符合写入测试用例的条件')
                    # func_box.通知机器人(f'脏数据过滤，这个不符合写入测试用例的条件 \n\n预期写入数据`{i}`\n\n```\n{test_case_content}\n```')
        # test_case
        sort_test_case = name_de_add_sort(test_case, sort_index)
        # 正式准备写入文件
        xlsx_heandle = ExcelHandle(ipaddr=llm_kwargs['ipaddr'], temp_file=template_file, sheet=sheet)
        xlsx_heandle.split_merged_cells()  # 先把合并的单元格拆分，避免写入失败
        file_path = xlsx_heandle.lpvoid_lpbuffe(sort_test_case, filename=long_name_processing(file_name))
        chat_file_list += f'{file_name}生成结果如下:\t {func_box.html_view_blank(__href=file_path, to_tabs=True)}\n\n'
        chatbot[-1] = ([you_say, chat_file_list])
        yield from toolbox.update_ui(chatbot, history)
        files_limit.append(file_path)
    return files_limit


def supplementary_test_case(gpt_response_collection, llm_kwargs, plugin_kwargs, chatbot, history):
    template_file, sheet, sort_index= json_args_return(plugin_kwargs, ['写入指定模版', '写入指定Sheet', '用例下标排序'])
    if not sheet:
        sheet, = json_args_return(plugin_kwargs, ['读取指定Sheet'])
    file_classification = file_classification_to_dict(gpt_response_collection)
    chat_file_list = ''
    you_say = '准备将测试用例写入Excel中...'
    chatbot.append([you_say, chat_file_list])
    yield from toolbox.update_ui(chatbot, history)
    files_limit = []
    for file_name in file_classification:
        old_file = plugin_kwargs['上阶段文件']
        old_case = plugin_kwargs[old_file]['原测试用例数据']
        header = plugin_kwargs[old_file]['原测试用例表头']
        test_case, desc = parsing_json_in_text(file_classification[file_name], old_case, filter_list=header, sort_index=sort_index)
        file_path = ExcelHandle(ipaddr=llm_kwargs['ipaddr'],
                                temp_file=template_file, sheet=sheet).lpvoid_lpbuffe(
            test_case, filename=long_name_processing(file_name))
        md = Utils().write_markdown(data=desc, hosts=llm_kwargs['ipaddr'], file_name=long_name_processing(file_name))
        chat_file_list += f'{file_name}生成结果如下:\t {func_box.html_view_blank(__href=file_path, to_tabs=True)}\n\n' \
                          f'---\n\n{file_name}补充思路如下：\t{func_box.html_view_blank(__href=md, to_tabs=True)}\n\n'
        chatbot[-1] = ([you_say, chat_file_list])
        yield from toolbox.update_ui(chatbot, history)
        files_limit.append(file_path)
    return files_limit


def transfer_flow_chart(gpt_response_collection, llm_kwargs, plugin_kwargs, chatbot, history):
    """
    Args: 将输出结果写入md，并转换为流程图
        gpt_response_collection: [输入、输出]
        llm_kwargs: 调优参数
        chatbot: 对话组件
        history: 历史对话
    Returns:
        None
    """
    file_classification = file_classification_to_dict(gpt_response_collection)
    file_limit = []
    chat_file_list = ''
    you_say = '请将Markdown结果转换为流程图~'
    chatbot.append([you_say, chat_file_list])
    for file_name in file_classification:
        inputs_count = ''
        for value in file_classification[file_name]:
            inputs_count += str(value).replace('```', '')  # 去除头部和尾部的代码块, 避免流程图堆在一块
        md, html = Utils().markdown_to_flow_chart(data=inputs_count, hosts=llm_kwargs['ipaddr'],
                                                  file_name=long_name_processing(file_name))
        chat_file_list += "View: " + func_box.html_view_blank(md,
                                                              to_tabs=True) + '\n\n--- \n\n View: ' + func_box.html_view_blank(
            html)
        chatbot.append((you_say, chat_file_list))
        yield from toolbox.update_ui(chatbot=chatbot, history=history, msg='成功写入文件！')
        file_limit.append(md)
    # f'tips: 双击空白处可以放大～\n\n' f'{func_box.html_iframe_code(html_file=html)}'  无用，不允许内嵌网页了
    return file_limit


def result_written_to_markdwon(gpt_response_collection, llm_kwargs, plugin_kwargs, chatbot, history, stage=''):
    """
    Args: 将输出结果写入md
        gpt_response_collection: [输入、输出]
        llm_kwargs: 调优参数
        chatbot: 对话组件
        history: 历史对话
    Returns:
        None
    """
    file_classification = file_classification_to_dict(gpt_response_collection)
    file_limit = []
    chat_file_list = ''
    you_say = '请将Markdown结果写入文件中...'
    chatbot.append([you_say, chat_file_list])
    for file_name in file_classification:
        inputs_all = ''
        for value in file_classification[file_name]:
            inputs_all += value
        md = Utils().write_markdown(data=inputs_all, hosts=llm_kwargs['ipaddr'],
                                    file_name=long_name_processing(file_name) + stage)
        chat_file_list = f'markdown已写入文件，下次使用插件可以直接提交markdown文件啦 {func_box.html_view_blank(md, to_tabs=True)}'
        chatbot[-1] = [you_say, chat_file_list]
        yield from toolbox.update_ui(chatbot=chatbot, history=history, msg='成功写入文件！')
        file_limit.append(md)
    return file_limit


# <---------------------------------------一些Tips----------------------------------------->
previously_on_plugins = f'如果是本地文件，请点击【🔗】先上传，多个文件请上传压缩包，' \
                        f'{func_box.html_tag_color("如果是网络文件或金山文档链接，请粘贴到输入框")}, 然后再次点击该插件' \
                        f'多个文件{func_box.html_tag_color("请使用换行或空格区分")}'

if __name__ == '__main__':
    test = [1, 2, 3, 4, [12], 33, 1]

    # with open('/Users/kilig/Desktop/test.md', mode='r') as f:
    #     content = f.read()
